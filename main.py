"""
微信支付账单处理器 - 手机版
功能：将Download/WeiXin文件夹中的Excel文件添加导出时间并保存到微信支付文件夹
"""

from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.progressbar import ProgressBar
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.clock import Clock
from kivy.utils import platform
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl.utils import get_column_letter

# Android权限处理
try:
    from android.permissions import request_permissions, Permission
    from android.storage import primary_external_storage_path
    ANDROID_AVAILABLE = True
except ImportError:
    ANDROID_AVAILABLE = False
    print("警告: android模块不可用，可能在非Android环境运行")

class ExcelProcessorApp(App):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.processed_count = 0
        self.total_count = 0
        self.current_time = None
        
    def build(self):
        # 获取存储路径
        self.source_folder = self.get_source_path()
        self.target_folder = self.get_target_path()
        
        # 请求权限
        if platform == 'android' and ANDROID_AVAILABLE:
            request_permissions([
                Permission.READ_EXTERNAL_STORAGE,
                Permission.WRITE_EXTERNAL_STORAGE
            ])
        
        # 创建主布局
        main_layout = BoxLayout(orientation='vertical', padding=15, spacing=10)
        
        # 标题
        title = Label(
            text='微信支付账单处理器',
            size_hint=(1, 0.1),
            font_size='20sp',
            bold=True,
            color=(0.2, 0.6, 1, 1)
        )
        main_layout.add_widget(title)
        
        # 状态信息区域（滚动视图）
        self.status_grid = GridLayout(cols=1, spacing=5, size_hint_y=None)
        self.status_grid.bind(minimum_height=self.status_grid.setter('height'))
        
        scroll_view = ScrollView(size_hint=(1, 0.5))
        scroll_view.add_widget(self.status_grid)
        main_layout.add_widget(scroll_view)
        
        # 添加初始状态信息
        self.add_status("程序初始化...")
        self.add_status(f"源文件夹: {self.source_folder}")
        self.add_status(f"目标文件夹: {self.target_folder}")
        
        # 进度条
        self.progress_bar = ProgressBar(max=100, value=0, size_hint=(1, 0.05))
        main_layout.add_widget(self.progress_bar)
        
        # 进度标签
        self.progress_label = Label(
            text='等待开始...',
            size_hint=(1, 0.05),
            font_size='14sp'
        )
        main_layout.add_widget(self.progress_label)
        
        # 按钮布局
        button_layout = BoxLayout(size_hint=(1, 0.15), spacing=10)
        
        # 开始处理按钮
        self.process_btn = Button(
            text='开始处理',
            background_color=(0.2, 0.6, 1, 1),
            font_size='18sp'
        )
        self.process_btn.bind(on_press=self.start_processing)
        button_layout.add_widget(self.process_btn)
        
        # 退出按钮
        exit_btn = Button(
            text='退出',
            background_color=(1, 0.3, 0.3, 1),
            font_size='18sp'
        )
        exit_btn.bind(on_press=self.stop)
        button_layout.add_widget(exit_btn)
        
        main_layout.add_widget(button_layout)
        
        # 启动时自动检查文件
        Clock.schedule_once(lambda dt: self.check_files(), 0.5)
        
        return main_layout
    
    def get_source_path(self):
        """获取源文件夹路径"""
        if platform == 'android' and ANDROID_AVAILABLE:
            primary_ext = primary_external_storage_path()
            return os.path.join(primary_ext, 'Download', 'WeiXin')
        else:
            # 测试环境
            return os.path.join(os.path.expanduser('~'), 'test_source')
    
    def get_target_path(self):
        """获取目标文件夹路径"""
        if platform == 'android' and ANDROID_AVAILABLE:
            primary_ext = primary_external_storage_path()
            return os.path.join(primary_ext, '微信支付')
        else:
            # 测试环境
            return os.path.join(os.path.expanduser('~'), 'test_target')
    
    def add_status(self, message):
        """添加状态信息"""
        label = Label(
            text=message,
            size_hint_y=None,
            height=30,
            halign='left',
            text_size=(self.width, None),
            font_size='14sp'
        )
        self.status_grid.add_widget(label)
        # 滚动到底部
        Clock.schedule_once(lambda dt: self.root.ids.scroll_view.scroll_to(label), 0.1)
    
    def check_files(self):
        """检查源文件夹中的文件"""
        self.add_status("正在检测文件...")
        
        try:
            if not os.path.exists(self.source_folder):
                self.add_status(f"错误：源文件夹不存在 - {self.source_folder}")
                self.add_status("请在手机存储的 Download/WeiXin 文件夹中放入Excel文件")
                return
            
            excel_files = [f for f in os.listdir(self.source_folder) 
                          if f.endswith(('.xlsx', '.xls'))]
            
            if excel_files:
                self.total_count = len(excel_files)
                self.add_status(f"找到 {self.total_count} 个Excel文件")
                self.progress_label.text = f"准备处理 {self.total_count} 个文件"
            else:
                self.add_status("未找到Excel文件！")
                self.add_status("请将Excel文件放入 Download/WeiXin 文件夹")
                
        except Exception as e:
            self.add_status(f"检测文件时出错: {str(e)}")
    
    def start_processing(self, instance):
        """开始处理文件"""
        # 禁用按钮
        self.process_btn.disabled = True
        self.process_btn.text = '处理中...'
        self.process_btn.background_color = (0.5, 0.5, 0.5, 1)
        
        # 清空状态信息
        self.status_grid.clear_widgets()
        self.add_status("开始处理文件...")
        
        # 异步执行处理
        Clock.schedule_once(lambda dt: self.process_files(), 0.1)
    
    def process_files(self):
        """处理文件的主函数"""
        try:
            # 检查源文件夹
            if not os.path.exists(self.source_folder):
                self.add_status(f"错误：源文件夹不存在")
                self.reset_button()
                return
            
            # 创建目标文件夹
            if not os.path.exists(self.target_folder):
                os.makedirs(self.target_folder)
                self.add_status(f"创建目标文件夹: {self.target_folder}")
            
            # 获取Excel文件
            excel_files = [f for f in os.listdir(self.source_folder) 
                          if f.endswith(('.xlsx', '.xls'))]
            
            if not excel_files:
                self.add_status("没有找到Excel文件！")
                self.reset_button()
                return
            
            self.total_count = len(excel_files)
            self.add_status(f"开始处理 {self.total_count} 个Excel文件")
            
            # 初始时间：当前系统时间
            self.current_time = datetime.now()
            
            # 处理每个文件
            self.processed_count = 0
            
            for i, filename in enumerate(excel_files):
                self.process_single_file(filename, i)
            
            # 处理完成
            self.add_status(f"\n处理完成！成功处理 {self.processed_count}/{self.total_count} 个文件")
            self.add_status(f"文件保存在: {self.target_folder}")
            self.progress_label.text = "处理完成！"
            
        except Exception as e:
            self.add_status(f"处理过程中出错: {str(e)}")
        
        finally:
            self.reset_button()
            # 重新检测文件
            Clock.schedule_once(lambda dt: self.check_files(), 1)
    
    def process_single_file(self, filename, index):
        """处理单个文件"""
        source_path = os.path.join(self.source_folder, filename)
        
        try:
            # 格式化时间
            time_str = self.current_time.strftime("%Y-%m-%d %H:%M:%S")
            time_str_for_filename = self.current_time.strftime("%Y%m%d%H%M%S")
            
            self.add_status(f"[{index+1}/{self.total_count}] 处理: {filename}")
            self.add_status(f"  添加时间: {time_str}")
            
            # 加载Excel文件
            wb = openpyxl.load_workbook(source_path)
            ws = wb.active
            
            # 在A5单元格添加导出时间
            ws['A5'] = f"导出时间：[{time_str}]"
            
            # 生成新文件名
            name_without_ext = os.path.splitext(filename)[0]
            ext = os.path.splitext(filename)[1]
            new_filename = f"{name_without_ext}_{time_str_for_filename}{ext}"
            target_path = os.path.join(self.target_folder, new_filename)
            
            # 保存文件
            wb.save(target_path)
            wb.close()
            
            self.add_status(f"  保存为: {new_filename}")
            
            # 时间增加45秒
            self.current_time += timedelta(seconds=45)
            self.processed_count += 1
            
        except Exception as e:
            self.add_status(f"  处理失败: {str(e)}")
        
        # 更新进度条
        progress = int(((index + 1) / self.total_count) * 100)
        self.progress_bar.value = progress
        self.progress_label.text = f"进度: {index+1}/{self.total_count} ({progress}%)"
    
    def reset_button(self):
        """重置按钮状态"""
        self.process_btn.disabled = False
        self.process_btn.text = '开始处理'
        self.process_btn.background_color = (0.2, 0.6, 1, 1)

if __name__ == '__main__':
    ExcelProcessorApp().run()